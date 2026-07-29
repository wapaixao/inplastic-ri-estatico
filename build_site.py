from __future__ import annotations

import json
import math
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path('/root/data/clientes/grupo-inplastic/apresentacao/Grupo_Inplastic_BP_DRE_2025_PREVIA_BALANCETES.xlsx')
OUT = Path('/root/workspace/inplastic-ri-estatico')


def clean(v):
    if v is None:
        return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    return v


def sheet_to_records(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c).strip() if c is not None else f'Coluna {i+1}' for i, c in enumerate(rows[0])]
    records = []
    for row in rows[1:]:
        if not any(c is not None for c in row):
            continue
        rec = {headers[i]: clean(row[i]) if i < len(row) else None for i in range(len(headers))}
        records.append(rec)
    return records


def num(v):
    return float(v or 0)


def find(records, label):
    for r in records:
        if str(r.get('Descrição', '')).strip().lower() == label.lower():
            return r
    return None

wb = load_workbook(SOURCE, data_only=True, read_only=True)
bp = sheet_to_records(wb, 'BP Consolidado')
dre = sheet_to_records(wb, 'DRE Consolidada')
inventario = sheet_to_records(wb, 'Inventário PDFs')
validacoes = sheet_to_records(wb, 'Validações')
leia = sheet_to_records(wb, 'Leia-me')

companies = ['Inplastic', 'Taoplast', 'Licitaplas', 'Plas Capital']

rl = find(dre, 'Receita Líquida / Resultado Receita') or {}
lucro = find(dre, 'Lucro Líquido Estimado') or find(dre, 'Lucro Líquido') or dre[-1]
ativo = find(bp, 'ATIVO') or {}
passivo = find(bp, 'PASSIVO + PL') or find(bp, 'PASSIVO') or {}

cards = [
    {'label': 'Ativo total grupo', 'value': num(ativo.get('Total Grupo')), 'kind': 'currency'},
    {'label': 'Passivo total grupo', 'value': num(passivo.get('Total Grupo')), 'kind': 'currency'},
    {'label': 'Receita líquida grupo', 'value': num(rl.get('Total Grupo')), 'kind': 'currency'},
    {'label': 'Lucro líquido estimado', 'value': num(lucro.get('Total Grupo')), 'kind': 'currency'},
]

data = {
    'meta': {
        'title': 'InPlastic — RI / Resultados Gerenciais',
        'group': 'Grupo InPlastic',
        'period': 'Dezembro/2025',
        'source_file': str(SOURCE),
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'notice': 'Prévia automática a partir de PDFs de balancete. Revisar contas críticas e eventuais eliminações intercompany antes de uso externo.',
        'scope_note': 'Este site não inclui PIS/COFINS nem módulo de Lucros/Distribuição para este grupo.',
    },
    'companies': companies,
    'cards': cards,
    'bp': bp,
    'dre': dre,
    'inventory': inventario,
    'validations': validacoes,
    'readme': leia,
}

(OUT / 'data.json').write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
print('data.json gerado com', len(bp), 'linhas BP e', len(dre), 'linhas DRE')
